
from openpyxl import Workbook
from openpyxl.styles import Font
wb = Workbook()
sheet = wb.active
sheet.title = "Language"
wb.create_sheet(title="Capital")

lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code = ['KA', 'TS', 'TN']

sheet['A1'].value = "State"
sheet['B1'].value = "Language"
sheet['C1'].value = "Code"

ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft

for i in range(2, 5):
    sheet.cell(i,1).value = state[i - 2]
    sheet.cell(i,2).value = lang[i - 2]
    sheet.cell(i,3).value = code[i - 2]

wb.save("demo.xlsx")

sheet = wb["Capital"]

sheet['A1'].value = "State"
sheet['B1'].value = "Capital"
sheet['C1'].value = "Code"
ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft

for i in range(2, 5):
    sheet.cell(i,1).value = state[i - 2]
    sheet.cell(i,2).value = capital[i - 2]
    sheet.cell(i,3).value = code[i - 2]

wb.save("demo.xlsx")

srchCode = input("Enter state code for finding capital ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding capital for code", srchCode, "is", sheet.cell(row=i, column=2).value)

sheet = wb["Language"]

srchCode = input("Enter state code for finding language ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding language for code", srchCode, "is", sheet.cell(row=i, column=2).value)

wb.close()